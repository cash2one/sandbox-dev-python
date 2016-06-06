#!/usr/bin/python
#encoding:utf-8

import xlrd

workbook = xlrd.open_workbook('e:/myexcel.xls')
worksheets = workbook.sheet_names()
worksheet1 = workbook.sheet_by_name(worksheets[0])
#worksheet1 = workbook.sheets()[0]
#worksheet1 = workbook.sheet_by_index(0)

num_rows = worksheet1.nrows ##获取行数
for curr_row in range(num_rows):
    row = worksheet1.row_values(curr_row)
    print('row%s is %s' %(curr_row,row))

num_cols = worksheet1.ncols  ##获取列数
for curr_col in range(num_cols):
    col = worksheet1.col_values(curr_col)
    print('col%s is %s' %(curr_col,col))

for rown in range(num_rows):
    for coln in range(num_cols):
        print worksheet1.cell_value(rown,coln)
        print worksheet1.cell(rown,coln).value
        print worksheet1.row(rown)[coln].value
        print worksheet1.col(coln)[rown].value
        #获取单元格中值的类型，类型 0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error
        print worksheet1.cell_type(rown,coln)

		
import xlwt
f=r'E:/test.xls'
workbook = xlwt.Workbook()
sheet1 = workbook.add_sheet('sheet1',cell_overwrite_ok=True)
sheet2 = workbook.add_sheet('sheet2',cell_overwrite_ok=True)
sheet1.write(0,0,'this should overwrite1')
sheet1.write(0,1,'aaaaaaaaaaaa')
sheet2.write(0,0,'this should overwrite2')
sheet2.write(1,2,'bbbbbbbbbbbbb')

style = xlwt.XFStyle() #初始化样式
font = xlwt.Font() #为样式创建字体
font.name = 'Times New Roman'
font.bold = True #设置样式的字体
style.font = font
sheet2.write(0,1,'some bold Times text', style) #使用样式，最后一个参数

workbook.save(f) #有文件存在则直接覆盖


import xlrd
import xlutils.copy
f=r'e:/myexcel.xls'
rb = xlrd.open_workbook(f)
wb = xlutils.copy.copy(rb)
ws = wb.get_sheet(0) #通过sheet_by_index()获取的sheet对象没有write()方法
ws.write(1, 1, 'changed!')
wb.add_sheet('sheetnnn2',cell_overwrite_ok=True)
wb.save(f)


from pyExcelerator import *
w=Workbook()  #创建工作簿
ws=w.add_sheet('Sheet1')  #添加工作表
ws.write(0, 0, 'A1')  #写入单元格
w.save('book.xls')

import pyExcelerator
sheets = pyExcelerator.parse_xls('book.xls')
print sheets


from openpyxl import Workbook
wb = Workbook()
ws = wb.active    ##获取active的sheet
ws['A1'] = 42   ##直接对指定cell进行数据赋值
ws.append([1, 2, 3])  ##追加一行数据
import datetime
ws['A2'] = datetime.datetime.now()   ##支持python格式自动转换
wb.save("sample.xlsx")

from openpyxl import load_workbook
wb = load_workbook(filename=r'e:/myexcel.xlsx')
sheetnames = wb.get_sheet_names()
ws = wb.get_sheet_by_name(sheetnames[0]) #打开sheet1
for rx in range(1,ws.max_row+1): #对第一个column遍历打印值
   print ws.cell(row=rx,column=1).value
ws.append([2,3,4])
wb.save('e:/myexcel.xlsx'))



